import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

DB_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "database.db")
OUT_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    f"survey_export_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
)

TEAL      = "005C8A"
WHITE     = "FFFFFF"
ALT_ROW   = "EEF6FB"
BORDER_C  = "C8DDE8"

SHEET_COLORS = {
    "General Info":          "005C8A",
    "Resources":             "1A6B9A",
    "Expenses":              "2E7DAD",
    "Unit Details":          "007A5E",
    "CEMS Installations":    "1B6CA8",
    "CEMS Drift General":    "2478B5",
    "CEMS Calibration Cyl":  "2D84C2",
    "CEMS Drift Details":    "3690CE",
    "CEMS Calib PM":         "3F9BDB",
    "CEMS Calib Gas":        "48A6E8",
    "CEMS Maintenance":      "51B1F5",
    "CEMS Impl Level":       "5ABCFF",
    "CEQMS Installations":   "006B3C",
    "CEQMS Val General":     "1A7B4C",
    "CEQMS Val Details":     "2E8B5C",
    "CEQMS Maintenance":     "3C9B6C",
    "CEQMS Impl Level":      "4AAB7C",
    "CAAQMS Installations":  "7B4A00",
    "CAAQMS Drift General":  "8B5A10",
    "CAAQMS Calib Cyl":      "9B6A20",
    "CAAQMS Drift Details":  "AB7A30",
    "CAAQMS Maintenance":    "BB8A40",
    "CAAQMS Impl Level":     "CB9A50",
    "Challenges":            "6B3A8B",
    "Improvements":          "84C225",
}


def read(conn, table):
    try:
        return pd.read_sql_query(f"SELECT * FROM [{table}]", conn)
    except Exception:
        return pd.DataFrame()


def clean_cols(df):
    """Pretty-print column names."""
    df.columns = [c.replace("_", " ").title() for c in df.columns]
    return df


def join_industry(df, industry, id_col="industry_id"):
    """Prepend Submission ID and Industry Name, then drop the id column."""
    if df.empty or industry.empty:
        return df
    ref = industry[["id", "industry_name"]].copy()
    ref = ref.rename(columns={"id": id_col, "industry_name": "Industry Name"})
    ref["Submission ID"] = ref[id_col].apply(lambda x: f"CEEW-{int(x):04d}")
    ref = ref[["Submission ID", "Industry Name", id_col]]
    merged = ref.merge(df, on=id_col, how="right")
    merged = merged.drop(columns=[id_col], errors="ignore")
    return merged


def style_sheet(ws, header_color=TEAL):
    thin  = Side(style="thin", color=BORDER_C)
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill("solid", fgColor=header_color)
    afill = PatternFill("solid", fgColor=ALT_ROW)
    wfill = PatternFill("solid", fgColor=WHITE)

    for cell in ws[1]:
        cell.font      = Font(bold=True, color=WHITE, name="Calibri", size=9)
        cell.fill      = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = bdr
    ws.row_dimensions[1].height = 36

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = afill if row_idx % 2 == 0 else wfill
        for cell in row:
            cell.fill      = fill
            cell.font      = Font(name="Calibri", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = bdr

    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        header = str(col_cells[0].value or "")
        if any(k in header for k in ("Id", "Year", "ID")):
            width = 14
        elif any(k in header for k in ("Name", "Address", "Responsibilit",
                                        "Activities", "Explanation", "Query",
                                        "Challenge", "Suggestion")):
            width = 38
        else:
            max_len = max(
                (len(str(c.value)) for c in col_cells[1:] if c.value),
                default=0
            )
            width = max(16, min(max_len + 2, 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "C2"   # freeze Submission ID + Industry Name columns



print("Connecting to database …")
conn = sqlite3.connect(DB_PATH)

industry = read(conn, "industry_profiles")
if industry.empty:
    print("No submissions found in database.")
    conn.close()
    raise SystemExit(0)

print(f"  {len(industry)} submission(s) found.\n")

SHEETS = [
    ("General Info",         "industry_profiles",              False),
    ("Resources",            "resource_availability",          True),
    ("Expenses",             "expenses_monitoring",            True),
    ("Unit Details",         "unit_details",                   True),
    ("CEMS Installations",   "cems_installations",             True),
    ("CEMS Drift General",   "cems_drift_general",             True),
    ("CEMS Calibration Cyl", "cems_calibration_cylinders",     True),
    ("CEMS Drift Details",   "cems_drift_details",             True),
    ("CEMS Calib PM",        "cems_calibration_pm",            True),
    ("CEMS Calib Gas",       "cems_calibration_gas",           True),
    ("CEMS Maintenance",     "cems_maintenance",               True),
    ("CEMS Impl Level",      "cems_implementation_levels",     True),
    ("CEQMS Installations",  "ceqms_installations",            True),
    ("CEQMS Val General",    "ceqms_validation_general",       True),
    ("CEQMS Val Details",    "ceqms_validation_details",       True),
    ("CEQMS Maintenance",    "ceqms_maintenance",              True),
    ("CEQMS Impl Level",     "ceqms_implementation_levels",    True),
    ("CAAQMS Installations", "caaqms_installations",           True),
    ("CAAQMS Drift General", "caaqms_drift_general",           True),
    ("CAAQMS Calib Cyl",     "caaqms_calibration_cylinders",   True),
    ("CAAQMS Drift Details", "caaqms_drift_details",           True),
    ("CAAQMS Maintenance",   "caaqms_maintenance",             True),
    ("CAAQMS Impl Level",    "caaqms_implementation_levels",   True),
    ("Challenges",           "implementation_challenges",      True),
    ("Improvements",         "expected_improvements",          True),
]

print("Writing Excel …")
written_sheets = []

with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
    pd.DataFrame([["Generating…"]]).to_excel(writer, sheet_name="_placeholder", index=False, header=False)

    for sheet_name, table_name, do_join in SHEETS:
        df = read(conn, table_name)
        if df.empty:
            print(f"  [skip] {sheet_name} — no data")
            continue

        if do_join:
            df = df.drop(columns=["id"], errors="ignore")
            df = join_industry(df, industry, id_col="industry_id")
        else:
            df["Submission ID"] = df["id"].apply(lambda x: f"CEEW-{int(x):04d}")
            df = df.drop(columns=["id"], errors="ignore")
            cols = ["Submission ID"] + [c for c in df.columns if c != "Submission ID"]
            df = df[cols]

        df = clean_cols(df)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        written_sheets.append(sheet_name)
        print(f"  ✓  {sheet_name:<24}  {len(df)} row(s),  {len(df.columns)} col(s)")

    wb_tmp = writer.book
    if "_placeholder" in wb_tmp.sheetnames and written_sheets:
        del wb_tmp["_placeholder"]

conn.close()

print("\nApplying formatting …")
wb = load_workbook(OUT_PATH)
for sheet_name in wb.sheetnames:
    color = SHEET_COLORS.get(sheet_name, TEAL)
    style_sheet(wb[sheet_name], header_color=color)

wb.save(OUT_PATH)

print(f"\n  Saved to:\n    {OUT_PATH}")
